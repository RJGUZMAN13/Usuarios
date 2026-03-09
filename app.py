# -*- coding: utf-8 -*-
"""
Sistema Gestión de Usuarios - PUREM Industrial
"""

import streamlit as st
import pandas as pd
import firebase_admin
from firebase_admin import credentials, firestore
import datetime
import time
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------- INIT FIRESTORE ----------------
if "db" not in st.session_state:
    if not firebase_admin._apps:
        firebase_config = dict(st.secrets["firebase"])
        cred = credentials.Certificate(firebase_config)
        firebase_admin.initialize_app(cred)
    st.session_state.db = firestore.client()

db = st.session_state.db

# ---------------- CONFIG ----------------
st.set_page_config(
    page_title="Gestión de Usuarios",
    page_icon="logoo.png",
    layout="wide"
)

# ---------------- ESTILOS ----------------
st.markdown("""
<style>
html, body, .main {
    background-color: #0E1117;
    color: white;
    font-family: 'Montserrat', sans-serif;
}
/* MEJORA LOGIN CELULAR */
@media (max-width: 800px) {
    .login-banner {
        display: none !important;
    }
    .stColumn {
        width: 100% !important;
    }
}
.card {
    background-color: #1b1f2a;
    padding: 1.2em;
    border-radius: 12px;
    margin-bottom: 1em;
    transition: 0.3s ease;
}
.card:hover {
    transform: translateY(-5px);
    box-shadow: 0 8px 20px rgba(0,255,153,0.2);
}
.badge-tecnico {
    background-color: #009966;
    padding: 4px 10px;
    border-radius: 6px;
    font-size: 0.8rem;
}
.badge-supervisor {
    background-color: #0055aa;
    padding: 4px 10px;
    border-radius: 6px;
    font-size: 0.8rem;
}
.badge-admin {
    background-color: #003366;
    padding: 4px 10px;
    border-radius: 6px;
    font-size: 0.8rem;
}
.stButton>button {
    background: linear-gradient(135deg, #00ff99 0%, #009966 100%);
    border-radius: 8px;
    font-weight: 600;
    transition: 0.3s ease;
}
.stButton>button:hover {
    transform: translateY(-3px);
    box-shadow: 0 8px 20px rgba(0,255,153,0.4);
}
.footer {
    position: fixed;
    bottom: 10px;
    right: 20px;
    color: rgba(255,255,255,0.5);
    font-size: 0.9rem;
}
</style>
""", unsafe_allow_html=True)

# ---------------- SESSION ----------------
if "auth" not in st.session_state:
    st.session_state.auth = False
if "user" not in st.session_state:
    st.session_state.user = None
if "log_df" not in st.session_state:
    st.session_state.log_df = pd.DataFrame()
if "excel_buffer" not in st.session_state:
    st.session_state.excel_buffer = None

# ---------------- LOGIN ----------------
if not st.session_state.auth:
    col1, col2 = st.columns([2,2])
    with col1:
        st.markdown("""
        <div class="login-banner" style="background: linear-gradient(135deg, #001a0d 0%, #00331a 100%);
        height:100vh;padding:8% 5%;display:flex;flex-direction:column;justify-content:center;">
        <h1 style="font-size:4rem;">Purem by Eberspächer<br>
        <span style="color:#00ff99;">Mantenimiento Industrial</span></h1>
        <p style="border-left:3px solid #00ff99;padding-left:20px;">
        Plataforma para Gestión de Usuarios de Mantenimiento.</p>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        st.markdown("## INICIAR SESIÓN")
        mx_input = st.text_input("ID de Usuario (MX)")
        password_input = st.text_input("Contraseña", type="password")
        if st.button("Acceder"):
            doc = db.collection("empleados").document(mx_input.upper()).get()
            if doc.exists and doc.to_dict().get("password") == password_input:
                st.session_state.auth = True
                st.session_state.user = doc.to_dict()
                st.rerun()
            else:
                st.error("Credenciales incorrectas")

# ---------------- PANEL ADMIN ----------------
if st.session_state.auth:
    with st.sidebar:
        st.image("logoo.png", width=120)
        st.markdown(f"### {st.session_state.user['nombre']}")
        if st.button("Cerrar sesión"):
            st.session_state.auth = False
            st.rerun()

    st.title("👥 Panel de Administración")
    tab1, tab_manual, tab2, tab3 = st.tabs(["📤 Alta de usuarios", "➕ Alta Manual", "📋 Usuarios registrados", "📜 Historial de altas"])

    # ---------------- TAB 1: ALTA EXCEL ----------------
    with tab1:
        uploaded_file = st.file_uploader("Subir archivo Excel", type=["xlsx"])
        if uploaded_file:
            df = pd.read_excel(uploaded_file)
            st.info(f"{len(df)} registros detectados")
            if st.button("Procesar archivo"):
                progress = st.progress(0)
                resultados = []
                total = len(df)
                for i, row in df.iterrows():
                    mx_id = str(row["mx"]).upper()
                    doc_ref = db.collection("empleados").document(mx_id)
                    doc = doc_ref.get()
                    estado = "NO MODIFICADO"
                    if not doc.exists:
                        doc_ref.set({
                            "mx": mx_id,
                            "nombre": row["nombre"],
                            "unidad": row["unidad"],
                            "business_unit": row["business_unit"],
                            "emp_no": int(row["emp_no"]),
                            "password": str(row["password"]),
                            "role": row.get("role", "tecnico"),
                            "area": row.get("area", "General"),
                            "last_login": None,
                            "mantener_sesion": False
                        })
                        estado = "REGISTRADO"
                    resultados.append({
                        "MX": mx_id,
                        "Nombre": row["nombre"],
                        "Estado": estado,
                        "Procesado por": st.session_state.user["nombre"],
                        "Fecha": datetime.datetime.now().strftime("%d-%m-%Y %H:%M")
                    })
                    progress.progress((i+1)/total)
                    time.sleep(0.02)
                st.session_state.log_df = pd.DataFrame(resultados)

                # Crear Excel con formato profesional
                wb = Workbook()
                ws = wb.active
                ws.title = "Reporte Oficial"
                ws.merge_cells("A1:E1")
                ws["A1"] = "REPORTE OFICIAL DE CARGA DE USUARIOS - PUREM - PLANTA RAMOS ARIZPE"
                ws["A1"].font = Font(size=14, bold=True)
                ws["A1"].alignment = Alignment(horizontal="center")
                ws.append([])
                ws.append(["MX", "Nombre", "Estado", "Procesado por", "Fecha"])
                for col in range(1, 6):
                    cell = ws.cell(row=3, column=col)
                    cell.fill = PatternFill(start_color="00331A", end_color="00331A", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                row_start = 4
                for i, row in st.session_state.log_df.iterrows():
                    ws.append(list(row))
                    estado_cell = ws.cell(row=row_start+i, column=3)
                    if row["Estado"] == "REGISTRADO":
                        estado_cell.fill = PatternFill(start_color="00CC66", end_color="00CC66", fill_type="solid")
                    else:
                        estado_cell.fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max_length + 4
                buffer = BytesIO()
                wb.save(buffer)
                excel_data = buffer.getvalue()
                st.session_state.excel_buffer = excel_data

                db.collection("historial_altas").add({
                    "admin": st.session_state.user["nombre"],
                    "fecha": datetime.datetime.now().strftime("%d-%m-%Y %H:%M"),
                    "excel_content": excel_data,
                    "registros": len(resultados)
                })
                st.success("Proceso completado")

        if not st.session_state.log_df.empty:
            st.markdown("### 📊 Resultado de carga")
            st.dataframe(st.session_state.log_df, use_container_width=True)
            st.download_button("📥 Descargar reporte", st.session_state.excel_buffer, file_name="reporte_alta.xlsx")

    # ---------------- TAB MANUAL (CORREGIDO SIN ERROR) ----------------
    with tab_manual:
        st.subheader("Alta Manual de Usuario")
        # clear_on_submit=True limpia el formulario automáticamente al confirmar
        with st.form("manual_form", clear_on_submit=True):
            col_m1, col_m2 = st.columns(2)
            with col_m1:
                m_mx = st.text_input("MX").upper()
                m_nombre = st.text_input("Nombre")
                m_unidad = st.text_input("Unidad")
                m_bu = st.text_input("Business Unit")
                m_emp = st.number_input("Emp No", step=1)
            with col_m2:
                m_pass = st.text_input("Password", type="password")
                m_role = st.selectbox("Role", ["tecnico", "supervisor", "admin"])
                st.text_input("Área", value="General", disabled=True)
                st.checkbox("Mantener Sesión", value=False, disabled=True)
                st.text_input("Last Login", value="None", disabled=True)
            
            c_btn1, c_btn2 = st.columns([1,4])
            with c_btn1:
                confirmar = st.form_submit_button("Confirmar Alta")
            with c_btn2:
                cancelar = st.form_submit_button("Cancelar")

            if confirmar:
                if m_mx and m_nombre and m_pass:
                    # 1. Guardar en Firestore
                    db.collection("empleados").document(m_mx).set({
                        "mx": m_mx, "nombre": m_nombre, "unidad": m_unidad,
                        "business_unit": m_bu, "emp_no": m_emp, "password": m_pass,
                        "role": m_role, "area": "General", "last_login": None, "mantener_sesion": False
                    })
                    
                    # 2. Generar Excel para historial
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Reporte Oficial"
                    ws.merge_cells("A1:E1")
                    ws["A1"] = "REPORTE OFICIAL DE CARGA DE USUARIOS - PUREM - PLANTA RAMOS ARIZPE"
                    ws["A1"].font = Font(size=14, bold=True)
                    ws["A1"].alignment = Alignment(horizontal="center")
                    ws.append([])
                    ws.append(["MX", "Nombre", "Estado", "Procesado por", "Fecha"])
                    for col in range(1, 6):
                        cell = ws.cell(row=3, column=col)
                        cell.fill = PatternFill(start_color="00331A", end_color="00331A", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    
                    fecha_now = datetime.datetime.now().strftime("%d-%m-%Y %H:%M")
                    ws.append([m_mx, m_nombre, "REGISTRADO", st.session_state.user["nombre"], fecha_now])
                    ws.cell(row=4, column=3).fill = PatternFill(start_color="00CC66", end_color="00CC66", fill_type="solid")
                    
                    buffer = BytesIO()
                    wb.save(buffer)
                    
                    # 3. Guardar en Historial
                    db.collection("historial_altas").add({
                        "admin": st.session_state.user["nombre"],
                        "fecha": fecha_now,
                        "excel_content": buffer.getvalue(),
                        "target_mx": m_mx,
                        "target_nombre": m_nombre
                    })
                    
                    st.success(f"Usuario {m_mx} registrado. Formulario limpio.")
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error("Llena los campos obligatorios (MX, Nombre, Password)")
            
            if cancelar:
                st.rerun()

    # ---------------- TAB 2: REGISTRADOS ----------------
    with tab2:
        st.subheader("Usuarios actualmente en Firestore")
        docs = db.collection("empleados").stream()
        usuarios = [{"mx": d.id, **d.to_dict()} for d in docs]

        total_tecnicos = len([u for u in usuarios if u.get("role") == "tecnico"])
        total_supervisores = len([u for u in usuarios if u.get("role") == "supervisor"])
        total_admins = len([u for u in usuarios if u.get("role") == "admin"])

        colA, colB, colC = st.columns(3)
        colA.markdown(f"""<div style="background-color:#001f3f;padding:20px;border-radius:10px;text-align:center;"><h3 style="color:#00ff99;">🛠 Técnicos</h3><p style="font-size:2rem;font-weight:bold;">{total_tecnicos}</p></div>""", unsafe_allow_html=True)
        colB.markdown(f"""<div style="background-color:#001f3f;padding:20px;border-radius:10px;text-align:center;"><h3 style="color:#00ff99;">🧑‍💼 Supervisores</h3><p style="font-size:2rem;font-weight:bold;">{total_supervisores}</p></div>""", unsafe_allow_html=True)
        colC.markdown(f"""<div style="background-color:#001f3f;padding:20px;border-radius:10px;text-align:center;"><h3 style="color:#00ff99;">👑 Admins</h3><p style="font-size:2rem;font-weight:bold;">{total_admins}</p></div>""", unsafe_allow_html=True)

        search = st.text_input("🔎 Buscar por MX o Nombre")
        if search:
            usuarios = [u for u in usuarios if search.lower() in u["mx"].lower() or search.lower() in u["nombre"].lower()]

        tecnicos = [u for u in usuarios if u.get("role") == "tecnico"]
        supervisores = [u for u in usuarios if u.get("role") == "supervisor"]

        tab_tec, tab_sup = st.tabs(["🛠 Técnicos", "🧑‍💼 Supervisores"])

        def mostrar_usuario(u):
            badge = "badge-tecnico" if u.get("role") == "tecnico" else "badge-supervisor" if u.get("role") == "supervisor" else "badge-admin"
            confirm_key = f"confirm_delete_{u['mx']}"
            st.markdown(f"""<div class="card"><h4 style="color:#00ff99;">{u['nombre']} ({u['mx']}) <span class="{badge}">{u.get('role')}</span></h4><p><b>Unidad:</b> {u.get('unidad','-')} | <b>Área:</b> {u.get('area','-')}</p></div>""", unsafe_allow_html=True)
            if not st.session_state.get(confirm_key, False):
                if st.button("⋮ Opciones", key=f"opt_{u['mx']}"):
                    st.session_state[confirm_key] = True
                    st.rerun()
            if st.session_state.get(confirm_key, False):
                st.warning("¿Eliminar definitivamente?")
                cA, cB = st.columns(2)
                if cA.button("Confirmar", key=f"yes_{u['mx']}"):
                    db.collection("empleados").document(u["mx"]).delete()
                    del st.session_state[confirm_key]
                    st.rerun()
                if cB.button("Cancelar", key=f"no_{u['mx']}"):
                    del st.session_state[confirm_key]
                    st.rerun()

        with tab_tec:
            for u in tecnicos: mostrar_usuario(u)
        with tab_sup:
            for u in supervisores: mostrar_usuario(u)

    # ---------------- TAB 3: HISTORIAL ----------------
    with tab3:
        st.subheader("📜 Historial de altas realizadas")
        hist_docs = db.collection("historial_altas").order_by("fecha", direction=firestore.Query.DESCENDING).stream()
        
        found = False
        for doc in hist_docs:
            found = True
            h = doc.to_dict()
            st.markdown(f"""<div class="card"><h4 style="color:#00ff99;">Alta realizada por {h.get('admin', 'Desconocido')}</h4><p><b>Fecha:</b> {h.get('fecha', 'Sin fecha')}</p><p><b>MX:</b> {h.get('target_mx', '-') } | <b>Nombre:</b> {h.get('target_nombre','-')}</p></div>""", unsafe_allow_html=True)
            if "excel_content" in h:
                st.download_button("📥 Descargar reporte", h["excel_content"], file_name=f"reporte_{h.get('fecha','archivo')}.xlsx", key=doc.id)
        if not found:
            st.info("No hay historial registrado aún.")

# ---------------- FOOTER ----------------
st.markdown("""<div class="footer">DEVELOPED BY: JUAN RODRIGO GUZMÁN MARTÍNEZ</div>""", unsafe_allow_html=True)